#!/usr/bin/python
# coding=utf-8
'''
将长文本划分为短文本，并按行存放在新xls中
'''
__author__ = 'Paul'
import xlrd,xlwt
import re
import openpyxl



if __name__ == '__main__':
    #example
    xlsfile= 'E:\\ctdata_yxbx.xls'
    newfile='E:\\newyxbx.xlsx'
    sheetname = 'ctdata'
    newxlr=openpyxl.Workbook()
    newSheet=newxlr.create_sheet(sheetname)

    myxlr=xlrd.open_workbook(xlsfile)
    mySheet=myxlr.sheet_by_name(sheetname)
    nrows=mySheet.nrows
    k=0
    n = 1
    for num in range(nrows):
        testcase = mySheet.cell(num, 0).value
        
        list1 = re.split("。|；", testcase)[0:-1]

        for i in list1:
            if(i !=''):
                #newSheet.write(k, 0, i)
                #newSheet.write(k, 1, n)
                i.replace(' ','') 
                i.replace('\t','')
                i.strip()
                newSheet.cell(row = k+1,column =1).value=i
                newSheet.cell(row = k+1,column =2).value=n
                k = k + 1
        n=n+1

    newxlr.save(newfile)
